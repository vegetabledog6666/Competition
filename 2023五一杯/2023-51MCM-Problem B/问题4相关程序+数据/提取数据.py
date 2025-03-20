import matplotlib.pyplot as plt
import pandas as pd
from copy import deepcopy
from openpyxl import Workbook
plt.rcParams['font.sans-serif']=['SimHei']
plt.rcParams['axes.unicode_minus'] = False


data=pd.read_excel("C:\\Users\\Administrator\\Desktop\\vs工作界面\\附件2.xlsx")

tim=data.loc[:,"日期(年/月/日) (Date Y/M/D)"]
sta=data.loc[:,"发货城市 (Delivering city)"]
end=data.loc[:,"收货城市 (Receiving city)"]
cos=data.loc[:,"快递运输数量(件) (Express delivery quantity (PCS))"]

for i in range(23,28):
    wb=Workbook()
    ws=wb.active
    ws.cell(1,1).value="日期(年/月/日) (Date Y/M/D)"
    ws.cell(1,2).value="发货城市 (Delivering city)"
    ws.cell(1,3).value="收货城市 (Receiving city)"
    ws.cell(1,4).value="快递运输数量(件) (Express delivery quantity (PCS))"
    line=2
    j=0
    while(tim[j]=="2023-4-{}".format(i)):
        ws.cell(line,1).value=tim[j]
        ws.cell(line,2).value=sta[j]
        ws.cell(line,3).value=end[j]
        ws.cell(line,4).value=cos[j]
        j+=1
        line+=1
    wb.save("{}日数据.xlsx".format(i))