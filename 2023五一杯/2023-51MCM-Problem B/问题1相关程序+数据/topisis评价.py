import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook
plt.rcParams['font.family'] = ['sans-serif']
plt.rcParams['font.sans-serif'] = ['SimHei']

df=pd.read_excel("C:\\Users\\Administrator\\Desktop\\vs工作界面\\各个指标整理.xlsx")
df=df.dropna(axis=0,how="all")
data=df.values
##提取城市标签
X=data[:,:1]
##提取城市评价指标
data=data[:,1:]

##进行标准化
data=data/np.sum(data*data,axis=0)**0.5
max_score=np.max(data,axis=0)
min_score=np.min(data,axis=0)
##求出其对应与最大值与最小值的距离
max_dist=np.sum((max_score-data)*(max_score-data),axis=1)**0.5
min_dist=np.sum((min_score-data)*(min_score-data),axis=1)**0.5

##进行得分计算
final_score=(min_dist/(max_dist+min_dist))
final_score/=np.sum(final_score)

final_score=final_score.tolist()
X=X.tolist()

x=dict()
j=0
##将其变成一个字典，方便之后从大到小排序
for i in X:
    x[i[0]]=final_score[j]
    j+=1
res=sorted(x.items(),key=lambda x:x[1],reverse=True)
x,final_score=[],[]



##以下皆为输出+可视化
wb=Workbook()
ws=wb.active

ws.cell(1,1).value="城市"
ws.cell(1,2).value="得分"

for num,sco in res:
    x.append(num)
    final_score.append(sco)


for i in range(len(x)):
    ws.cell(i+2,1).value=x[i]
    ws.cell(i+2,2).value=final_score[i]
wb.save("topsis评分表.xlsx")

ax=plt.bar(x=x,height=final_score)##绘制条形图
plt.title("各个城市topsis评分排序")
plt.xlabel("各个城市")
plt.ylabel("各个城市得分情况")
plt.show()
