import xlrd
from openpyxl import Workbook
import numpy as np
import networkx as nx
import matplotlib.pyplot as plt

wb=xlrd.open_workbook("C:\\Users\\Administrator\\Desktop\\vs工作界面\\附件1.xlsx")
ws=wb.sheet_by_index(0)
left,right,num=[],[],[]
left=ws.col_values(1)
right=ws.col_values(2)
num=ws.col_values(3)
left.pop(0);right.pop(0);num.pop(0)
##上述代码为读入数据，然后去掉标签


jud=dict()

g=nx.DiGraph()##创建一个图

for i in range(0,len(num)):
    ##以下则是判断相应的点或边是否在图中，没有则加入
    if(left[i] not in jud.keys()):
        jud[left[i]]=1
        g.add_node(left[i])
    if(right[i] not in jud.keys()):
        jud[left[i]]=1
        g.add_node(right[i])
    if(g.has_edge(left[i],right[i])==0):
        g.add_edge(left[i],right[i],weight=num[i])
    else :
        g.edges[left[i],right[i]].update({"weight":num[i]+g.get_edge_data(left[i],right[i])["weight"]})

nx.draw_networkx(g,pos=nx.shell_layout(g),node_size=200,node_shape='o',width=1,style='solid',font_size=8)##进行可视化
plt.show()

des=Workbook()
ws=des.active
ws.cell(1,1).value="节点"
ws.cell(1,2).value="度中心值排序"
j=2

np=nx.pagerank(g,alpha=0.9)
top_k = sorted(np.items(), key=lambda x: x[1], reverse=True)
for node, value in top_k:
    ws.cell(j,1).value=node
    ws.cell(j,2).value=value
    j+=1
des.save("中心值的值.xlsx")