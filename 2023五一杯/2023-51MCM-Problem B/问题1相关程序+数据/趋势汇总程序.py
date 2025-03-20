import pandas as pd
import numpy as np
from openpyxl import Workbook
data=pd.read_excel("C:\\Users\\Administrator\\Desktop\\vs工作界面\\附件1.xlsx",parse_dates=["日期(年/月/日) (Date Y/M/D)"])

month=data['日期(年/月/日) (Date Y/M/D)'].dt.strftime('%Y-%m')
month_ori=month
##下列代码用来计算不重复的时间列表
month=month.drop_duplicates()

month.index=range(0,10)

nowline,nextline=0,0
needdata=Workbook()
des=needdata.active

##代表对其进行从头到尾遍历
for i in range(0,9):
    ##nowcity表示现在城市的快递收发货情况，nextcity表示下一个时间段城市的收发货情况，change表示相比于上个月的差异平均差异
    nowcity,nextcity,change=dict(),dict(),dict()
    ##下列两个变量分别用来判断该城市有没有进入过上述字典中
    jud1,jud2=dict(),dict()
    ##nowline表示现在遍历到的年份，只要符合现在的就可以继续往下遍历，一直遍历到下一年终止
    while(month_ori.iloc[nowline]==month[i]):
        bgi,end,num=data.iloc[nowline,1],data.iloc[nowline,2],data.iloc[nowline,3]
        if(bgi not in jud1.keys()):
            nowcity[bgi]=num
            jud1[bgi]=1
        else :
            nowcity[bgi]+=num
        if(end not in jud1.keys()):
            nowcity[end]=num
            jud1[end]=1
        else :
            nowcity[end]+=num
        nowline+=1
    
    ##然后遍历下一年的，同上，nextline表示其遍历到的位置
    nextline=nowline
    while(month_ori.iloc[nextline]==month[i+1]):
        bgi,end,num=data.iloc[nextline,1],data.iloc[nextline,2],data.iloc[nextline,3]
        if(bgi not in jud2.keys()):
            nextcity[bgi]=num
            jud2[bgi]=1
        else :
            nextcity[bgi]+=num
        if(end not in jud2.keys()):
            nextcity[end]=num
            jud2[end]=1
        else :
            nextcity[end]+=num
        nextline+=1
        if(nextline==16962):    ##表明这是数据的最后一行
            break
    
    same=list(set(list(nextcity.keys())+list(nowcity.keys())))##表明当前月份出现的城市列
    
    temp=2
    for j in sorted(same):
        if((j in nowcity.keys())&(j in nextcity.keys())):  ##如果都出现了，就计算其增长率
            change[j]=(nextcity[j]-nowcity[j])/nowcity[j]    ##计算某月的增加或下降趋势，后面两个是为了避免0
        elif(j not in nowcity.keys()):    ##如果上一年没有出现过，就赋值为1，否则为0
            change[j]=1        
        else :
            change[j]=-1
        des.cell(1+i*2,temp).value=j
        temp+=1
        
    des.cell(2+i*2,1).value=month.iloc[i+1]
    temp=1
    for j in sorted(same):
        des.cell(2+i*2,temp+1).value=change[j]
        temp+=1
needdata.save("趋势数据.xlsx")