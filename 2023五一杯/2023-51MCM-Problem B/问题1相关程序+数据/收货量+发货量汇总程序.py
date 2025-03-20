import xlrd
import numpy as np
from openpyxl import Workbook

wb=xlrd.open_workbook("C:\\Users\\Administrator\\Desktop\\vs工作界面\\附件1.xlsx")
ws=wb.sheet_by_index(0)
start,end,num=[],[],[]

##去掉头部的标签
start=ws.col_values(1)  
end=ws.col_values(2)
num=ws.col_values(3)
start.pop(0);end.pop(0);num.pop(0)


##deli为发货量，rece为收货量
deli=dict()
rece=dict()
for i in range(0,len(start)):
    ##以下两行代码的意思是如果城市号不在对应的字典的键中，那么就说明是第一次读到，进行赋值，否则加上即可
    if(start[i] not in deli.keys()):
        deli[start[i]]=num[i]
    else :
        deli[start[i]]+=num[i]
    if(end[i] not in rece.keys()):
        rece[end[i]]=num[i]
    else :
        rece[end[i]]+=num[i]

##以下存储发货量
res1=Workbook()
ws1=res1.active

##城市标签行赋值
ws1.cell(1,1).value="城市"
ws1.cell(1,2).value="发货量"

##以下逐行进行赋值
j=2
for name,tot in zip(list(deli.keys()),list(deli.values())):
    ws1.cell(j,1).value=name
    ws1.cell(j,2).value=tot
    j+=1
res1.save("发货量总结表.xlsx")

##以下为存储收货量
res2=Workbook()
ws2=res2.active
ws2.cell(1,1).value="城市"
ws2.cell(1,2).value="收货量"
j=2
for name,tot in zip(list(rece.keys()),list(rece.values())):
    ws2.cell(j,1).value=name
    ws2.cell(j,2).value=tot
    j+=1
res2.save("收货量总结表.xlsx")
