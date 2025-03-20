import pandas as pd
import numpy as np
from openpyxl import Workbook
data=pd.read_excel("C:\\Users\\Administrator\\Desktop\\vs工作界面\\附件1.xlsx",parse_dates=["日期(年/月/日) (Date Y/M/D)"])

tim=data['日期(年/月/日) (Date Y/M/D)'].dt.strftime("%Y/%m/%d")
tim_ori=tim
tim=tim.drop_duplicates()
tim=tim.values

b=input("请输入你要查找数据的路径起点：")
e=input("请输入你要查找数据的路径终点：")

nowline=0
needdata=Workbook()
des=needdata.active
daydata=dict()

for i in range(0,len(tim)):
    
    while(tim_ori.iloc[nowline]==tim[i]):
        bgi,end,num=data.iloc[nowline,1],data.iloc[nowline,2],data.iloc[nowline,3]
        if(bgi==b) & (end==e):
            if(tim[i] not in daydata.keys()):
                daydata[tim[i]]=num
            else :
                daydata[tim[i]]+=num
        nowline+=1
        if(nowline==16962):
            break
        
wb=Workbook()
ws=wb.active
ws.cell(1,1).value="时间"
ws.cell(1,2).value="总快递运输数量"
i=2
for temp1,temp2 in zip(list(daydata.keys()),list(daydata.values())):
    ws.cell(i,1).value=temp1
    ws.cell(i,2).value=temp2
    i+=1
wb.save("各天{}城到{}城快递运输数量提取收据.xlsx".format(b,e))