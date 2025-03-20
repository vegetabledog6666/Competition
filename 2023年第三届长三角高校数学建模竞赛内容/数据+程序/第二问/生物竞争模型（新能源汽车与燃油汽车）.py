import numpy as np
import matplotlib.pyplot  as plt
import xlrd
from openpyxl import Workbook
import sympy
from sklearn import preprocessing
plt.rcParams['font.family'] = ['sans-serif']
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus']=False

wk=xlrd.open_workbook(r"C:\Users\Administrator\Desktop\vs工作界面\生物竞争预测用数据.xlsx")
ws=wk.sheet_by_index(0)
x1=np.array(ws.col_values(1)[1:])
x2=np.array(ws.col_values(2)[1:])    ##市场占比的变化数据
r1=np.array(ws.col_values(4)[1:])    ##汽车保有量增长率
r2=np.array(ws.col_values(3)[1:])

scaler=preprocessing.MinMaxScaler(feature_range=(0.1,0.9))    ##防止数据变成0后面出现计算错误,将其映射到0.1与0.9上
r1=scaler.fit_transform(r1.reshape(-1,1))
r2=scaler.fit_transform(r2.reshape(-1,1))
r1=r1.reshape(1,-1)[0]    ##将数组推回到一维数组格式
r2=r2.reshape(1,-1)[0]

t=np.arange(2015,2023,1)

x1=np.round(x1,3)
x2=np.round(x2,3)      ##保留三位小数
r1=np.round(r1,3)
r2=np.round(r2,3)


x=sympy.Symbol("x")     ##新能源汽车的数据代入sigmoid函数运算
dx2dt=1/(1+sympy.exp(-x))   ##新能源汽车
dx1dt=1-1/(1+sympy.exp(-x))   ##燃油汽车
dx2dt=sympy.diff(dx2dt,x)     ##求对应的微分
dx1dt=sympy.diff(dx1dt,x)
dx2=[]
dx1=[]
for i in t:      ##求出对应的值
    dx2.append(float(dx2dt.evalf(subs={x:i})))
    dx1.append(float(dx1dt.evalf(subs={x:i})))
dx1=np.array(dx1)
dx2=np.array(dx2)
a1=(r1*x1*(1-x1)-dx1)/(x1*x2)
a2=(r2*x2*(1-x2)-dx2)/(x1*x2)
a1=np.round(a1,3)
a2=np.round(a2,3)


wb=Workbook()            ##导出
ws=wb.active
ws.cell(1,1).value="燃油汽车"
ws.cell(1,2).value="新能源汽车"
j=2
for x,y in zip(a1,a2):
    ws.cell(j,1).value=x
    ws.cell(j,2).value=y
    j+=1
wb.save("竞争系数表.xlsx")


plt.figure(figsize=(7,5))     ##画图
plt.scatter(t,a1,c="b",marker="o")
plt.scatter(t,a2,c="b",marker="*")
plt.plot(t,a1,color="r",linestyle=":",label="燃油汽车的竞争系数a1变化")
plt.plot(t,a2,color="r",linestyle="-",label="新能源汽车的竞争系数a2变化")
for x,y in zip(t,a1):
    plt.text(x,y-0.05,s="x={}\ny={}".format(x,y))
for x,y in zip(t,a2):
    plt.text(x,y,s="x={}\ny={}".format(x,y))
plt.title("燃油汽车与新能源汽车竞争系数随时间变化情况")
plt.legend(loc='upper right',fontsize=12)
plt.show()