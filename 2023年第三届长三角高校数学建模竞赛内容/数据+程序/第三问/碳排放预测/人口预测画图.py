import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook

x=np.arange(2023,2061,1)

y=224120000/(1/1.129+0.1234*0.9254**(x-2012))

des=Workbook()
ws=des.active
ws.cell(1,1).value="预测年份"
ws.cell(1,2).value="人口数量"
j=2
for a,b in zip(x,y):
    ws.cell(j,1).value=a
    ws.cell(j,2).value=b
    j+=1
des.save("人口预测表.xlsx")

plt.plot(x,y,c="r",linestyle="-")
plt.scatter(x,y,c="b",marker="*")
plt.show()