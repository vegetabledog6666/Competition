import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook

x=np.arange(2022,2061,1)

y=63045.13332143490/(1/5.758+0.9304*0.9048**(x-2012))

des=Workbook()
ws=des.active
ws.cell(1,1).value="预测年份"
ws.cell(1,2).value="人均GDP"
j=2
for a,b in zip(x,y):
    ws.cell(j,1).value=a
    ws.cell(j,2).value=b
    j+=1
des.save("人均GDP预测表.xlsx")

plt.plot(x,y,c="r",linestyle="-")
plt.scatter(x,y,c="b",marker="*")
plt.show()
