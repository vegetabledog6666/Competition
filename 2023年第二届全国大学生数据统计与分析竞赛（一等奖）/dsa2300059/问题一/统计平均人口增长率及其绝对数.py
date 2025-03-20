from openpyxl import load_workbook
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import math
plt.rcParams['font.family'] = ['sans-serif']
plt.rcParams['font.sans-serif'] = ['SimHei']


wb=load_workbook(r"C:\Users\Administrator\Desktop\vs工作界面\各个国家人口数据（已处理）.xlsx")
ws=wb.active

nation_growth=dict()
nation_num=dict()
sub_year=range(2011,2022)

for i in range(1,16057,1):    ##逐行读入，只统计2011年到2022年的，同时的话用字典判别不同国家
    year=ws.cell(i,2).value
    nation=ws.cell(i,1).value
    if(nation not in nation_growth.keys()):
        nation_growth[nation]=0
        nation_num[nation]=0
    if(year in sub_year):
        nation_growth[nation]+=ws.cell(i,4).value
        nation_num[nation]+=ws.cell(i,5).value

for i in nation_growth.keys():
    nation_growth[i]/=11
    nation_num[i]/=11
##对于人口增长率以及人口增长进行排序
res1=sorted(nation_growth.items(),key=lambda d:d[1],reverse=True)
res2=sorted(nation_num.items(),key=lambda d:d[1],reverse=True)


x,y=[],[]
print("人口增长率前10:")
for i in range(0,10):
    print(res1[i][0])
    x.append(res1[i][0])
    y.append(round(res1[i][1],4))
data=pd.DataFrame(data=np.transpose(np.array([x,y])),columns=["国家","人口增长率前10"])
data.to_excel("人口增长率前10表.xlsx")

x,y=[],[]
print("人口增长前10:")
for i in range(0,10):
    print(res2[i][0])
    x.append(res2[i][0])
    y.append(math.floor(res2[i][1]))
data=pd.DataFrame(data=np.transpose(np.array([x,y])),columns=["国家","人口增长前10"])
data.to_excel("人口增长前10表.xlsx")

x,y=[],[]
print("人口增长率倒数后10")
for i in reversed(res1[-10:]):
    print(i[0])
    x.append(i[0])
    y.append(round(-i[1],4))
data=pd.DataFrame(data=np.transpose(np.array([x,y])),columns=["国家","人口增长率倒数后10"])
data.to_excel("人口增长率倒数后10表.xlsx")

x,y=[],[]
print("人口增长倒数后10")
for i in reversed(res2[-10:]):
    print(i[0])
    x.append(i[0])
    y.append(math.floor(-i[1]))
data=pd.DataFrame(data=np.transpose(np.array([x,y])),columns=["国家","人口增长倒数后10"])
data.to_excel("人口增长倒数后10表.xlsx")