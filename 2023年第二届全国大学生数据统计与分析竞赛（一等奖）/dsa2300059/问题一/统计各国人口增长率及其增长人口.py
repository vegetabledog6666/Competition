from openpyxl import load_workbook

wb=load_workbook(r"C:\Users\Administrator\Desktop\vs工作界面\各个国家人口数据.xlsx")
ws=wb.active
ws.cell(1,4,"人口增长率")
ws.cell(1,5,"人口增长")

for i in range(2,15987,72):    ##就行数进行遍历，同时72是从一个国家跳到另外一个国家的步数，表示1950-2022这几年
    ws.cell(i,4,0)
    ws.cell(i,5,0)
    for j in range(1,72):
        ws.cell(i+j,5,ws.cell(i+j,3).value-ws.cell(i+j-1,3).value)
        ws.cell(i+j,4,ws.cell(i+j,5).value/ws.cell(i+j,3).value)

wb.save(r"C:\Users\Administrator\Desktop\vs工作界面\各个国家人口数据（已处理）.xlsx")
